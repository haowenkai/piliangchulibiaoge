import openpyxl

def load_workbook(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet_names = workbook.sheetnames
    return workbook, sheet_names

def save_workbook(workbook, file_path):
    workbook.save(file_path) 