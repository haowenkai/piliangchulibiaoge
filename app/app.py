import os
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, Toplevel, Text
from datetime import datetime
from dotenv import load_dotenv
from app.ui import UI
import openpyxl

class App:
    def __init__(self, root):
        self.root = root
        self.filepath = None
        self.workbook = None
        self.ui = None

    def set_ui(self, ui):
        self.ui = ui

    def open_file(self):
        filepath = filedialog.askopenfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx;*.xlsm;*.xltx;*.xltm")]
        )
        if filepath:
            try:
                self.filepath = filepath
                self.workbook = openpyxl.load_workbook(filepath)
                sheet_names = self.workbook.sheetnames
                self.ui.update_file_label(os.path.basename(filepath))
                modified_time = datetime.fromtimestamp(os.path.getmtime(filepath)).strftime("%Y-%m-%d %H:%M:%S")
                self.ui.update_modified_time_label(modified_time)
                self.ui.populate_sheet_list(sheet_names)
            except Exception as e:
                messagebox.showerror("错误", f"打开文件错误: {e}")

    def read_sheet_data(self, sheet_name):
        if self.workbook and sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)

            # 显示数据在一个新的窗口
            data_window = Toplevel(self.root)
            data_window.title(f"{sheet_name} 数据")
            text_area = Text(data_window, wrap=tk.NONE)
            text_area.pack(expand=True, fill='both')
            for row in data:
                text_area.insert(tk.END, "\t".join(map(str, row)) + "\n")
            text_area.config(state=tk.DISABLED)
        else:
            messagebox.showerror("错误", "工作簿未打开或工作表不存在")

class ExcelRenamerApp:
    def __init__(self, root, excel_manager):
        # 初始化UI
        self.root = root
        self.root.title(os.getenv("APP_TITLE", "批量修改表格名称"))
        self.root.geometry(f"{os.getenv('APP_WIDTH', '800')}x{os.getenv('APP_HEIGHT', '600')}")

        # 接收传递的 ExcelManager 实例
        self.excel_manager = excel_manager

        # 初始化UI，传递 excel_manager 实例
        self.ui = UI(self.root, self.excel_manager)

        self.file_label = self.ui.file_label
        self.modified_time_label = self.ui.modified_time_label
        self.sheet_listbox = self.ui.sheet_listbox
        rename_button = self.ui.rename_button
        delete_button = self.ui.delete_button
        read_data_button = self.ui.read_data_button
        exit_button = self.ui.exit_button
        self.time_label = self.ui.time_label

        # 关联按钮事件
        rename_button.config(command=self.rename_sheet)
        delete_button.config(command=self.delete_selected)
        read_data_button.config(command=self.read_sheet_data)
        self.ui.open_file_button.config(command=self.open_file_dialog)
        exit_button.config(command=self.root.destroy)

    def open_file_dialog(self):
        filepath = filedialog.askopenfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx;*.xlsm;*.xltx;*.xltm")]
        )
        if filepath:
            try:
                self.excel_manager.load_workbook(filepath)
                filename = os.path.basename(filepath)
                modified_time = datetime.fromtimestamp(os.path.getmtime(filepath)).strftime("%Y-%m-%d %H:%M:%S")
                self.ui.update_file_label(filename)
                self.ui.update_modified_time_label(modified_time)
                self.ui.populate_sheet_list(self.excel_manager.workbook.sheetnames)
            except Exception as e:
                messagebox.showerror("错误", f"打开文件错误: {e}")

    def read_sheet_data(self):
        selected_sheet = self.ui.sheet_listbox.get(tk.ANCHOR)
        if selected_sheet:
            data = self.excel_manager.read_sheet_data(selected_sheet)
            # 显示数据在一个新的窗口
            data_window = Toplevel(self.root)
            data_window.title(f"{selected_sheet} 数据")
            text_area = Text(data_window, wrap=tk.NONE)
            text_area.pack(expand=True, fill='both')
            for row in data:
                text_area.insert(tk.END, "\t".join(map(str, row)) + "\n")
            text_area.config(state=tk.DISABLED)
        else:
            messagebox.showinfo("提示", "请选择要读取的工作表")

    def rename_sheet(self):
        selected_sheet = self.ui.sheet_listbox.get(tk.ANCHOR)
        if not selected_sheet:
            messagebox.showinfo("提示", "请选择要重命名的工作表")
            return

        new_name = simpledialog.askstring("重命名", f"将 '{selected_sheet}' 重命名为:")
        if new_name:
            try:
                self.excel_manager.rename_sheet(selected_sheet, new_name)
                self.ui.populate_sheet_list(self.excel_manager.workbook.sheetnames)
                messagebox.showinfo("成功", f"工作表 '{selected_sheet}' 已重命名为 '{new_name}'")
            except Exception as e:
                messagebox.showerror("错误", f"重命名工作表错误: {e}")

    def delete_selected(self):
        selected_sheet = self.ui.sheet_listbox.get(tk.ANCHOR)
        if selected_sheet:
            if messagebox.askyesno("确认", f"确定要删除工作表 '{selected_sheet}' 吗?"):
                try:
                    self.excel_manager.delete_sheet(selected_sheet)
                    self.ui.populate_sheet_list(self.excel_manager.workbook.sheetnames)
                    messagebox.showinfo("成功", f"工作表 '{selected_sheet}' 已删除")
                except Exception as e:
                    messagebox.showerror("错误", f"删除工作表错误: {e}")
        else:
            messagebox.showinfo("提示", "请选择要删除的工作表")
