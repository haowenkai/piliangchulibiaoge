from openpyxl import Workbook
from pathlib import Path
from datetime import datetime
import os
from typing import Tuple, List, Optional
from app.logging_config import logger
import openpyxl

class ExcelManager:
    def __init__(self):
        self.file_path: Optional[Path] = None
        self.workbook: Optional[Workbook] = None
        self.sheet_names: List[str] = []

    def load_workbook(self, file_path: str) -> Tuple[Workbook, List[str]]:
        """加载Excel文件"""
        try:
            self.file_path = Path(file_path)
            self.workbook = openpyxl.load_workbook(file_path)
            self.sheet_names = self.workbook.sheetnames
            logger.info(f"成功加载文件: {self.file_path}")
            return self.workbook, self.sheet_names
        except Exception as e:
            logger.error(f"加载文件失败: {str(e)}")
            raise

    def save_workbook(self) -> None:
        """保存Excel文件"""
        if self.workbook and self.file_path:
            try:
                self.workbook.save(self.file_path)
                logger.info(f"成功保存文件: {self.file_path}")
            except Exception as e:
                logger.error(f"保存文件失败: {str(e)}")
                raise

    def rename_sheet(self, old_name: str, new_name: str) -> None:
        """重命名工作表"""
        if self.workbook:
            try:
                if new_name in self.sheet_names and new_name != old_name:
                    raise ValueError(f"工作表名称 '{new_name}' 已存在")

                self.workbook[old_name].title = new_name
                index = self.sheet_names.index(old_name)
                self.sheet_names[index] = new_name
                logger.info(f"成功重命名工作表: {old_name} -> {new_name}")
            except Exception as e:
                logger.error(f"重命名工作表失败: {str(e)}")
                raise

    def delete_sheet(self, sheet_name: str) -> None:
        """删除工作表"""
        if self.workbook:
            try:
                del self.workbook[sheet_name]
                self.sheet_names.remove(sheet_name)
                logger.info(f"成功删除工作表: {sheet_name}")
            except Exception as e:
                logger.error(f"删除工作表失败: {str(e)}")
                raise

    def get_file_info(self) -> Tuple[str, str]:
        """获取文件信息"""
        if self.file_path and self.file_path.exists():
            filename = self.file_path.name
            modified_time = datetime.fromtimestamp(
                self.file_path.stat().st_mtime
            ).strftime('%Y-%m-%d %H:%M:%S')
            return filename, modified_time
        return "未打开", "未知"

    def read_sheet_data(self, sheet_name: str) -> List[List]:
        """读取工作表数据"""
        if self.workbook and sheet_name in self.sheet_names:
            sheet = self.workbook[sheet_name]
            data = []
            for row in sheet.iter_rows():
                row_data = [cell.value for cell in row]
                data.append(row_data)
            logger.info(f"成功读取工作表 '{sheet_name}' 的数据")
            return data
        else:
            logger.error(f"无法读取工作表 '{sheet_name}'，工作表不存在或工作簿未加载")
            return []

    def write_sheet_data(self, sheet_name: str, data: List[List]) -> None:
        """写入工作表数据"""
        if self.workbook and sheet_name in self.sheet_names:
            sheet = self.workbook[sheet_name]
            for row_index, row_data in enumerate(data):
                for col_index, cell_value in enumerate(row_data):
                    sheet.cell(row=row_index + 1, column=col_index + 1, value=cell_value)
            logger.info(f"成功向工作表 '{sheet_name}' 写入数据")
        else:
            logger.error(f"无法向工作表 '{sheet_name}' 写入数据，工作表不存在或工作簿未加载")

    def add_sheet(self, sheet_name: str) -> None:
        """添加新的工作表"""
        if self.workbook:
            try:
                self.workbook.create_sheet(sheet_name)
                self.sheet_names.append(sheet_name)
                logger.info(f"成功添加新的工作表: {sheet_name}")
            except Exception as e:
                logger.error(f"添加新的工作表失败: {str(e)}")
                raise

    def modify_cell_data(self, sheet_name: str, row: int, column: int, value) -> None:
        """修改单元格数据"""
        if self.workbook and sheet_name in self.sheet_names:
            sheet = self.workbook[sheet_name]
            try:
                sheet.cell(row=row, column=column).value = value
                logger.info(f"成功修改工作表 '{sheet_name}' 的单元格 ({row}, {column}) 的值为: {value}")
            except Exception as e:
                logger.error(f"修改工作表 '{sheet_name}' 的单元格 ({row}, {column}) 失败: {str(e)}")
        else:
            logger.error(f"无法修改工作表 '{sheet_name}' 的单元格，工作表不存在或工作簿未加载")
